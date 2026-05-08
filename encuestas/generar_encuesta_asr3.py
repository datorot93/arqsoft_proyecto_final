"""
Genera la encuesta de validación del ASR-3 (Facilidad de Integración) en formato Excel.

Salida: encuestas/encuesta_validacion_asr3.xlsx
- Hoja "Encuesta": 10 preguntas, peso, calificación 1-5, justificación, aporte ponderado.
- Hoja "Resultado": total ponderado, umbral, veredicto.
- Hoja "Escala": leyenda 1-5.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

PREGUNTAS = [
    {
        "id": "Q1",
        "dimension": "Estándares de comunicación",
        "pregunta": (
            "¿En qué medida los componentes ApiGateway, CDTXPais, CreditoExpressXPais y "
            "MotorElegibilidadXPais exponen contratos en estándares industriales formales "
            "(OpenAPI 3.x para REST, AsyncAPI para los tópicos del MessageBroker, "
            "gRPC/Protobuf para RPC interno) verificables a partir del diagrama de "
            "componentes y del diagrama de clases?"
        ),
        "componentes": "ApiGateway, CDTXPais, CreditoExpressXPais, MotorElegibilidadXPais, MessageBroker",
        "peso": 0.12,
    },
    {
        "id": "Q2",
        "dimension": "Anti-Corruption Layer (ACL) y desacople semántico",
        "pregunta": (
            "¿Qué tan claro queda el subsistema Integracion (AdaptadorCore + TraductorDominio + "
            "CircuitBreaker) como una Anti-Corruption Layer que evita la fuga del modelo legacy "
            "SOAP/CICS de CoreBancoZ hacia los servicios de dominio CDTXPais, CreditoExpressXPais "
            "y Pagos, manteniendo un modelo de dominio limpio en LineaVerde?"
        ),
        "componentes": "AdaptadorCore, TraductorDominio, CircuitBreaker, CoreBancoZ, CDTXPais, CreditoExpressXPais, Pagos",
        "peso": 0.15,
    },
    {
        "id": "Q3",
        "dimension": "Asincronía y desacople temporal con el Core",
        "pregunta": (
            "¿El uso del MessageBroker (tópicos cdt.eventos, credito.eventos, saldo.cambios, eventos) "
            "y de ChangeDataCapture reduce el acoplamiento síncrono con CoreBancoZ y permite que la "
            "elegibilidad de Crédito Express fluya event-driven sin esperar respuestas síncronas del core?"
        ),
        "componentes": "MessageBroker, ChangeDataCapture, CoreBancoZ, MotorElegibilidadXPais, ActualizadorCache",
        "peso": 0.12,
    },
    {
        "id": "Q4",
        "dimension": "Versionado de eventos y schemas",
        "pregunta": (
            "¿El diseño contempla un Schema Registry y políticas de evolución (compatibilidad backward/"
            "forward) para los eventos publicados en el MessageBroker, de modo que MotorElegibilidadXPais, "
            "Notificaciones, DetectorFraude y ActualizadorCache puedan evolucionar de manera independiente "
            "sin romper a productores existentes?"
        ),
        "componentes": "MessageBroker, MotorElegibilidadXPais, Notificaciones, DetectorFraude, ActualizadorCache",
        "peso": 0.10,
    },
    {
        "id": "Q5",
        "dimension": "Versionado de APIs públicas y compatibilidad hacia atrás",
        "pregunta": (
            "¿El ApiGateway formaliza una estrategia explícita de versionado de APIs públicas "
            "(URI versioning, header versioning o contract testing) que permita a AplicacionMovil "
            "y AplicacionWeb coexistir con múltiples versiones simultáneas durante despliegues "
            "azules/verdes o canary?"
        ),
        "componentes": "ApiGateway, AplicacionMovil, AplicacionWeb",
        "peso": 0.08,
    },
    {
        "id": "Q6",
        "dimension": "Integración con proveedores externos heterogéneos",
        "pregunta": (
            "¿El diseño especifica con suficiente claridad la integración con sistemas externos "
            "(ValidadorIdentidad, ConsultaProveedores y Convenios), incluyendo clientes dedicados, "
            "contratos definidos, mecanismos de timeout y fallback, de forma que un cambio en un "
            "proveedor externo no propague rotura a los demás?"
        ),
        "componentes": "ValidadorIdentidad, ConsultaProveedores, Convenios, OnboardingXPais, MotorElegibilidadXPais, Pagos",
        "peso": 0.10,
    },
    {
        "id": "Q7",
        "dimension": "Automatización de elegibilidad bajo el SLA de 10 minutos",
        "pregunta": (
            "¿La cadena CDTXPais → MessageBroker.cdt.eventos → MotorElegibilidadXPais "
            "(con consulta a ConsultaProveedores y materialización del resultado en CacheDistribuido) "
            "sustituye plenamente las validaciones manuales que hoy demoran 4 días, manteniéndose "
            "dentro del umbral del ASR-3 de elegibilidad < 10 minutos de forma automática?"
        ),
        "componentes": "CDTXPais, MessageBroker, MotorElegibilidadXPais, ConsultaProveedores, CacheDistribuido",
        "peso": 0.12,
    },
    {
        "id": "Q8",
        "dimension": "Cohesión interna y bajo acoplamiento entre subsistemas",
        "pregunta": (
            "¿Las dependencias en el diagrama de componentes muestran que el subsistema LineaVerde "
            "se acopla a CoreBancoZ exclusivamente a través del subsistema Integracion (sin "
            "atajos directos) y a Datos por almacén dedicado por contexto (AlmacenCDTXPais, "
            "AlmacenCreditoXPais, AlmacenPagosXPais), respetando el principio de bounded context?"
        ),
        "componentes": "LineaVerde, Integracion, CoreBancoZ, AlmacenCDTXPais, AlmacenCreditoXPais, AlmacenPagosXPais",
        "peso": 0.08,
    },
    {
        "id": "Q9",
        "dimension": "Capacidad de evolución multi-país (multi-tenancy)",
        "pregunta": (
            "¿El patrón XPais (sufijo en CDTXPais, CreditoExpressXPais, OnboardingXPais, "
            "AlmacenCDTXPais, AlmacenCreditoXPais, AlmacenPagosXPais y el atributo pais : Pais "
            "en el diagrama de clases) permite incorporar un nuevo país, o evolucionar las reglas "
            "de un país existente, sin modificar los componentes de los demás países?"
        ),
        "componentes": "CDTXPais, CreditoExpressXPais, OnboardingXPais, AlmacenCDTXPais, AlmacenCreditoXPais, AlmacenPagosXPais",
        "peso": 0.06,
    },
    {
        "id": "Q10",
        "dimension": "Resiliencia y aislamiento de fallos en la integración",
        "pregunta": (
            "¿El CircuitBreaker, los pools de hilos/conexiones acotados en el AdaptadorCore y el "
            "buffering del MessageBroker garantizan que una degradación de CoreBancoZ no se "
            "traduzca en fallos en cascada hacia CDTXPais, CreditoExpressXPais, Pagos y los canales "
            "(AplicacionMovil, AplicacionWeb)?"
        ),
        "componentes": "CircuitBreaker, AdaptadorCore, MessageBroker, CoreBancoZ, CDTXPais, CreditoExpressXPais, Pagos",
        "peso": 0.07,
    },
]

UMBRAL = 4.0

# --- Estilos ---
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
BORDER_THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
ALIGN_TOP_WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
ALIGN_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

wb = Workbook()

# ===== Hoja 1: Encuesta =====
ws = wb.active
ws.title = "Encuesta"

ws["A1"] = "Encuesta de Validación del ASR-3 — Facilidad de Integración"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")

ws["A2"] = (
    "Caso: Banco Z – Línea Verde · Apertura de CDT y elegibilidad para Crédito Express. "
    "Escala: 1=No cumple en absoluto, 2=Cumple parcialmente bajo, 3=Cumple parcialmente, "
    "4=Cumple en alto grado, 5=Cumple a cabalidad. Umbral de aprobación: promedio ponderado > 4.0."
)
ws["A2"].alignment = ALIGN_TOP_WRAP
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 50

# Encabezados (fila 4)
headers = [
    ("ID", 6),
    ("Dimensión técnica", 28),
    ("Pregunta", 70),
    ("Componentes a analizar (diagramas_final)", 35),
    ("Peso", 8),
    ("Calificación (1-5)", 12),
    ("Justificación del experto", 40),
    ("Aporte ponderado", 14),
]
for col_idx, (text, width) in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_THIN
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[4].height = 32

# Validación: calificación 1..5
dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
dv.error = "Ingrese un entero entre 1 y 5"
dv.errorTitle = "Calificación inválida"
dv.prompt = "1=No cumple, 5=Cumple a cabalidad"
dv.promptTitle = "Escala"
ws.add_data_validation(dv)

# Filas de preguntas (5..14)
for i, p in enumerate(PREGUNTAS):
    r = 5 + i
    ws.cell(row=r, column=1, value=p["id"]).alignment = ALIGN_CENTER
    ws.cell(row=r, column=2, value=p["dimension"]).alignment = ALIGN_TOP_WRAP
    ws.cell(row=r, column=3, value=p["pregunta"]).alignment = ALIGN_TOP_WRAP
    ws.cell(row=r, column=4, value=p["componentes"]).alignment = ALIGN_TOP_WRAP
    peso_cell = ws.cell(row=r, column=5, value=p["peso"])
    peso_cell.number_format = "0.00"
    peso_cell.alignment = ALIGN_CENTER
    cal_cell = ws.cell(row=r, column=6)  # vacía, para que la complete el experto
    cal_cell.alignment = ALIGN_CENTER
    dv.add(cal_cell)
    ws.cell(row=r, column=7).alignment = ALIGN_TOP_WRAP
    aporte_cell = ws.cell(row=r, column=8, value=f"=IFERROR(E{r}*F{r},\"\")")
    aporte_cell.number_format = "0.000"
    aporte_cell.alignment = ALIGN_CENTER
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = BORDER_THIN
    ws.row_dimensions[r].height = 110

# Fila de totales (16)
total_row = 5 + len(PREGUNTAS) + 1
ws.cell(row=total_row, column=2, value="TOTAL / VEREDICTO").font = Font(bold=True)
ws.cell(row=total_row, column=4, value="Suma de pesos:").alignment = Alignment(horizontal="right")
ws.cell(row=total_row, column=5, value=f"=SUM(E5:E{4+len(PREGUNTAS)})").number_format = "0.00"
ws.cell(row=total_row, column=7, value="Promedio ponderado:").alignment = Alignment(horizontal="right")
prom_cell = ws.cell(row=total_row, column=8, value=f"=IFERROR(SUM(H5:H{4+len(PREGUNTAS)})/SUM(E5:E{4+len(PREGUNTAS)}),\"\")")
prom_cell.number_format = "0.000"
prom_cell.font = Font(bold=True, size=12)
prom_cell.fill = PatternFill("solid", fgColor="FFF2CC")

# Veredicto (17)
ver_row = total_row + 1
ws.cell(row=ver_row, column=7, value="Veredicto:").alignment = Alignment(horizontal="right")
ws.cell(row=ver_row, column=7).font = Font(bold=True)
ver_cell = ws.cell(
    row=ver_row,
    column=8,
    value=f'=IF(H{total_row}="","Pendiente",IF(H{total_row}>{UMBRAL},"CUMPLIDO","NO CUMPLIDO"))',
)
ver_cell.font = Font(bold=True, size=12)
ver_cell.alignment = ALIGN_CENTER

# Formato condicional sobre el veredicto y el promedio
green_fill = PatternFill("solid", fgColor="C6EFCE")
red_fill = PatternFill("solid", fgColor="FFC7CE")
ws.conditional_formatting.add(
    f"H{ver_row}",
    CellIsRule(operator="equal", formula=['"CUMPLIDO"'], fill=green_fill),
)
ws.conditional_formatting.add(
    f"H{ver_row}",
    CellIsRule(operator="equal", formula=['"NO CUMPLIDO"'], fill=red_fill),
)
ws.conditional_formatting.add(
    f"H{total_row}",
    CellIsRule(operator="greaterThan", formula=[str(UMBRAL)], fill=green_fill),
)
ws.conditional_formatting.add(
    f"H{total_row}",
    CellIsRule(operator="lessThanOrEqual", formula=[str(UMBRAL)], fill=red_fill),
)

# Inmovilizar paneles
ws.freeze_panes = "A5"

# ===== Hoja 2: Escala =====
ws2 = wb.create_sheet("Escala")
ws2["A1"] = "Escala de calificación"
ws2["A1"].font = TITLE_FONT
escala = [
    (1, "No cumple en absoluto", "El componente o decisión no aparece o contradice la propiedad evaluada."),
    (2, "Cumple parcialmente bajo", "Aparece de forma incipiente; gaps significativos."),
    (3, "Cumple parcialmente", "Aparece pero con ambigüedad o decisiones pendientes."),
    (4, "Cumple en alto grado", "Decisión presente, justificada y trazable a los diagramas."),
    (5, "Cumple a cabalidad", "Decisión explícita, completa y consistente en componentes/clases/despliegue."),
]
ws2.cell(row=3, column=1, value="Valor").font = HEADER_FONT
ws2.cell(row=3, column=1).fill = HEADER_FILL
ws2.cell(row=3, column=2, value="Etiqueta").font = HEADER_FONT
ws2.cell(row=3, column=2).fill = HEADER_FILL
ws2.cell(row=3, column=3, value="Descripción").font = HEADER_FONT
ws2.cell(row=3, column=3).fill = HEADER_FILL
for i, (v, et, desc) in enumerate(escala):
    r = 4 + i
    ws2.cell(row=r, column=1, value=v).alignment = ALIGN_CENTER
    ws2.cell(row=r, column=2, value=et)
    ws2.cell(row=r, column=3, value=desc).alignment = ALIGN_TOP_WRAP
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 80

# ===== Hoja 3: Resultado =====
ws3 = wb.create_sheet("Resultado")
ws3["A1"] = "Resultado de la evaluación del ASR-3"
ws3["A1"].font = TITLE_FONT
ws3["A3"] = "Promedio ponderado:"
ws3["A3"].font = Font(bold=True)
ws3["B3"] = f"=Encuesta!H{total_row}"
ws3["B3"].number_format = "0.000"
ws3["A4"] = "Umbral de aprobación:"
ws3["A4"].font = Font(bold=True)
ws3["B4"] = f"> {UMBRAL}"
ws3["A5"] = "Veredicto:"
ws3["A5"].font = Font(bold=True)
ws3["B5"] = f"=Encuesta!H{ver_row}"
ws3["B5"].font = Font(bold=True, size=12)
ws3.conditional_formatting.add(
    "B5", CellIsRule(operator="equal", formula=['"CUMPLIDO"'], fill=green_fill)
)
ws3.conditional_formatting.add(
    "B5", CellIsRule(operator="equal", formula=['"NO CUMPLIDO"'], fill=red_fill)
)
ws3["A7"] = (
    "Cálculo: cada pregunta tiene un peso wi (sumando 1.00) y una calificación si en [1,5]. "
    "El promedio ponderado se calcula como Σ(wi · si) / Σ(wi). El ASR-3 se considera CUMPLIDO "
    "únicamente si el promedio ponderado es estrictamente > 4.0."
)
ws3["A7"].alignment = ALIGN_TOP_WRAP
ws3.merge_cells("A7:F9")
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 22

OUT = "/home/datorot/arqsoft_proyecto_final/encuestas/encuesta_validacion_asr3.xlsx"
wb.save(OUT)
print(f"Generado: {OUT}")
print(f"Suma de pesos: {sum(p['peso'] for p in PREGUNTAS):.2f}")
